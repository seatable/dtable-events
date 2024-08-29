import openpyxl
import os
import shutil
import uuid
from copy import deepcopy

from dtable_events.dtable_io.excel import parse_row, write_xls_with_type, TEMP_EXPORT_VIEW_DIR, IMAGE_TMP_DIR
from dtable_events.dtable_io.utils import get_related_nicknames_from_dtable, get_metadata_from_dtable_server, \
    escape_sheet_name
from dtable_events.utils import get_inner_dtable_server_url, get_location_tree_json, gen_random_option, format_date
from dtable_events.utils.constants import ColumnTypes
from dtable_events.app.config import INNER_DTABLE_DB_URL, BIG_DATA_ROW_IMPORT_LIMIT, BIG_DATA_ROW_UPDATE_LIMIT, \
    ARCHIVE_VIEW_EXPORT_ROW_LIMIT, APP_TABLE_EXPORT_EXCEL_ROW_LIMIT
from dtable_events.utils.dtable_db_api import DTableDBAPI, convert_db_rows
from dtable_events.utils.dtable_server_api import DTableServerAPI
from dtable_events.utils.sql_generator import filter2sql, BaseSQLGenerator

AUTO_GENERATED_COLUMNS = [
    ColumnTypes.AUTO_NUMBER,
    ColumnTypes.CTIME,
    ColumnTypes.MTIME,
    ColumnTypes.CREATOR,
    ColumnTypes.LAST_MODIFIER,
    ColumnTypes.BUTTON,
    ColumnTypes.FORMULA,
    ColumnTypes.LINK_FORMULA,
]

ROW_EXCEED_ERROR_CODE = 1
FILE_READ_ERROR_CODE = 2
COLUMN_MATCH_ERROR_CODE = 3
ROW_INSERT_ERROR_CODE = 4
INTERNAL_ERROR_CODE = 5


def _parse_excel_row(excel_row_data, column_name_type_map, name_to_email, location_tree, excel_select_column_options):
    parsed_row_data = {}
    for col_name, value in excel_row_data.items():
        col_type = column_name_type_map.get(col_name)
        if not value:
            continue
        excel_value = parse_row(col_type, value, name_to_email, location_tree=location_tree)

        if excel_value:
            if col_type in [ColumnTypes.SINGLE_SELECT, ColumnTypes.MULTIPLE_SELECT]:
                col_options = excel_select_column_options.get(col_name, set())
                if not col_options:
                    excel_select_column_options[col_name] = col_options
                if col_type == ColumnTypes.MULTIPLE_SELECT:
                    col_options.update(set(excel_value))
                else:
                    col_options.add(excel_value)

        parsed_row_data[col_name] = excel_value
    return parsed_row_data, excel_select_column_options


def handle_excel_row_datas(db_api, table_name, excel_row_datas, ref_cols, column_name_type_map, column_name_data_map, name_to_email, location_tree, insert_new_row=False):
    where_clauses = []
    for ref_col in ref_cols:
        value_list = []
        none_in_list = False
        for row_data in excel_row_datas:
            value = row_data.get(ref_col)
            if not value:
                none_in_list = True
            if value and value not in value_list:
                value_list.append(value)
        if none_in_list:
            where_clauses.append(
                "(`%s` in (%s) or `%s` is null)" % (
                    ref_col,
                    str(value_list).replace('[', '').replace(']', ''),
                    ref_col)
            )
        else:
            where_clauses.append(
                "`%s` in (%s)" % (
                    ref_col,
                    str(value_list).replace('[', '').replace(']', ''),
                )
            )

    sql = "Select * from `%s` where %s" % (
        table_name,
        ' And '.join(where_clauses)
    )

    rows_for_import = []
    rows_for_update = []

    query_rows_from_base, db_metadata = db_api.query(sql, convert=True, server_only=False)
    query_rows_from_base = convert_db_rows(db_metadata, query_rows_from_base)
    excel_select_column_options = {}
    for excel_row in excel_row_datas:
        excel_ref_data = {col: excel_row.get(col) for col in ref_cols if excel_row.get(col)}
        parsed_row, excel_select_column_options = _parse_excel_row(excel_row, column_name_type_map, name_to_email, location_tree, excel_select_column_options)

        find_tag = False
        for base_row in query_rows_from_base:
            base_ref_data = {}
            for col in ref_cols:
                if base_row.get(col):
                    base_ref_data[col] = base_row.get(col)
                    if column_name_type_map.get(col) and column_name_type_map.get(col) == ColumnTypes.DATE:
                        pre_format_date_str = base_row.get(col).replace('T', ' ')
                        pre_format_date_str = pre_format_date_str.split('+' if '+' in pre_format_date_str else '-')[0]
                        base_ref_data[col] = format_date(pre_format_date_str, column_name_data_map.get(col).get('format'))
            if base_ref_data and excel_ref_data and base_ref_data == excel_ref_data:
                rows_for_update.append({
                    "row_id": base_row.get('_id'),
                    "row": parsed_row  # parse
                })
                find_tag = True
        if insert_new_row and excel_ref_data and not find_tag:
            rows_for_import.append(parsed_row)  # parse
    return rows_for_import, rows_for_update, excel_select_column_options


def add_column_options(dtable_server_api, table_name, excel_select_column_options, dtable_col_name_to_column):
    # add select column options
    is_added_options = False
    for col_name, excel_options in excel_select_column_options.items():
        column = dtable_col_name_to_column.get(col_name)
        col_data = column.get('data')
        dtable_options = col_data and col_data.get('options') or []
        to_be_added_options = excel_options - set([op.get('name') for op in dtable_options])
        if to_be_added_options:
            options = [gen_random_option(option) for option in to_be_added_options]
            dtable_server_api.add_column_options(table_name, col_name, options)
            is_added_options = True
    return is_added_options

def import_excel_to_db(
        username,
        dtable_uuid,
        table_name,
        file_path,
        task_id,
        tasks_status_map

):
    from dtable_events.dtable_io import dtable_io_logger

    tasks_status_map[task_id] = {
        'status': 'initializing',
        'err_msg': '',
        'rows_imported': 0,
        'err_code': 0,
    }
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheets = wb.get_sheet_names()
        ws = wb[sheets[0]]
    except Exception as err:
        tasks_status_map[task_id]['err_msg'] = "file reading error: %s" % str(err)
        tasks_status_map[task_id]['status'] = 'terminated'
        tasks_status_map[task_id]['err_code'] = FILE_READ_ERROR_CODE
        os.remove(file_path)
        return

    try:

        dtable_server_url = get_inner_dtable_server_url()

        base = DTableServerAPI(username, dtable_uuid, dtable_server_url)
        base_columns = base.list_columns(table_name)
        column_name_type_map = {}
        column_name_data_map = {}
        for col in base_columns:
            column_name_type_map[col.get('name')] = col.get('type')
            column_name_data_map[col.get('name')] = col.get('data')
        matched_columns = []
        excel_columns = []
        for cell in ws[1]:
            excel_columns.append(cell.value)
            col_type = column_name_type_map.get(cell.value)
            if col_type and col_type not in AUTO_GENERATED_COLUMNS:
                matched_columns.append(cell.value)

        if not matched_columns:
            tasks_status_map[task_id]['err_msg'] = 'No matching columns in excel'
            tasks_status_map[task_id]['status'] = 'terminated'
            tasks_status_map[task_id]['err_code'] = COLUMN_MATCH_ERROR_CODE
            os.remove(file_path)
            return

        db_handler = DTableDBAPI(username, dtable_uuid, INNER_DTABLE_DB_URL)
    except Exception as err:
        tasks_status_map[task_id]['err_msg'] = str(err)
        tasks_status_map[task_id]['status'] = 'terminated'
        tasks_status_map[task_id]['err_code'] = INTERNAL_ERROR_CODE
        os.remove(file_path)
        return

    total_count = 0
    insert_count = 0
    slice_data = []

    status = 'success'
    tasks_status_map[task_id]['status'] = 'running'

    related_users = get_related_nicknames_from_dtable(dtable_uuid)
    name_to_email = {user.get('name'): user.get('email') for user in related_users}

    location_tree = get_location_tree_json()

    index = 0
    exceed_flag = False

    excel_select_column_options = {}
    dtable_col_name_to_column = {col['name']: col for col in base_columns}
    for row in ws.rows:
        if index > BIG_DATA_ROW_IMPORT_LIMIT:
            exceed_flag = True
            break
        try:
            if index > 0:  # skip header row
                row_list = [r.value for r in row]
                row_data = dict(zip(excel_columns, row_list))
                parsed_row_data = {}
                for col_name, value in row_data.items():
                    col_type = column_name_type_map.get(col_name)
                    col_data = column_name_data_map.get(col_name)
                    if not col_type or col_type in AUTO_GENERATED_COLUMNS:
                        continue

                    excel_value = value and parse_row(col_type, value, name_to_email, location_tree=location_tree, column_data=col_data) or ''
                    if excel_value:
                        if col_type in [ColumnTypes.SINGLE_SELECT, ColumnTypes.MULTIPLE_SELECT]:
                            col_options = excel_select_column_options.get(col_name, set())
                            if not col_options:
                                excel_select_column_options[col_name] = col_options
                            if col_type == ColumnTypes.MULTIPLE_SELECT:
                                col_options.update(set(excel_value))
                            else:
                                col_options.add(excel_value)
                    parsed_row_data[col_name] = excel_value

                slice_data.append(parsed_row_data)
                if len(slice_data) == 100:
                    tasks_status_map[task_id]['rows_imported'] = insert_count
                    is_added_options = add_column_options(base, table_name, excel_select_column_options, dtable_col_name_to_column)
                    if is_added_options:
                        # if column option is added, update base_columns
                        base_columns = base.list_columns(table_name)
                        dtable_col_name_to_column = {col['name']: col for col in base_columns}
                        excel_select_column_options = {}
                    db_handler.insert_rows(table_name, slice_data)
                    insert_count += len(slice_data)
                    slice_data = []
                total_count += 1
            index += 1
        except Exception as err:
            tasks_status_map[task_id]['err_msg'] = 'Row inserted error'
            tasks_status_map[task_id]['status'] = 'terminated'
            tasks_status_map[task_id]['err_code'] = ROW_INSERT_ERROR_CODE
            dtable_io_logger.exception(err)
            os.remove(file_path)
            return

    if slice_data:
        add_column_options(base, table_name, excel_select_column_options, dtable_col_name_to_column)
        db_handler.insert_rows(table_name, slice_data)
        insert_count += len(slice_data)

    if exceed_flag:
        tasks_status_map[task_id]['err_msg'] = 'Number of rows exceeds %s limit' % BIG_DATA_ROW_IMPORT_LIMIT
        tasks_status_map[task_id]['status'] = 'terminated'
        tasks_status_map[task_id]['err_code'] = ROW_EXCEED_ERROR_CODE
        tasks_status_map[task_id]['rows_imported'] = insert_count
        os.remove(file_path)
        return

    tasks_status_map[task_id]['status'] = status
    tasks_status_map[task_id]['rows_imported'] = insert_count
    os.remove(file_path)
    return


def update_excel_to_db(
        username,
        dtable_uuid,
        table_name,
        file_path,
        ref_columns,
        is_insert_new_data,
        task_id,
        tasks_status_map

):
    from dtable_events.dtable_io import dtable_io_logger

    tasks_status_map[task_id] = {
        'status': 'initializing',
        'err_msg': '',
        'rows_handled': 0,
        'err_code': 0,
    }
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheets = wb.get_sheet_names()
        ws = wb[sheets[0]]
    except Exception as err:
        tasks_status_map[task_id]['err_msg'] = "file reading error: %s" % str(err)
        tasks_status_map[task_id]['status'] = 'terminated'
        tasks_status_map[task_id]['err_code'] = FILE_READ_ERROR_CODE
        os.remove(file_path)
        return
    ref_columns = ref_columns.split(',')
    try:
        dtable_server_url = get_inner_dtable_server_url()
        excel_columns = [cell.value for cell in ws[1]]
        for ref_col in ref_columns:
            if ref_col not in excel_columns:
                tasks_status_map[task_id]['err_msg'] = 'Column %s does not exist in excel' % ref_col
                tasks_status_map[task_id]['status'] = 'terminated'
                tasks_status_map[task_id]['err_code'] = COLUMN_MATCH_ERROR_CODE
                os.remove(file_path)
                return

        db_handler = DTableDBAPI(username, dtable_uuid, INNER_DTABLE_DB_URL)
        base = DTableServerAPI(username, dtable_uuid, dtable_server_url)
        base_columns = base.list_columns(table_name)
        column_name_type_map = {}
        column_name_data_map = {}
        for col in base_columns:
            column_name_type_map[col.get('name')] = col.get('type')
            column_name_data_map[col.get('name')] = col.get('data')
    except Exception as err:
        tasks_status_map[task_id]['err_msg'] = str(err)
        tasks_status_map[task_id]['status'] = 'terminated'
        tasks_status_map[task_id]['err_code'] = INTERNAL_ERROR_CODE
        os.remove(file_path)
        return

    total_count = 0  # data in excel scanned
    related_users = get_related_nicknames_from_dtable(dtable_uuid)
    name_to_email = {user.get('name'): user.get('email') for user in related_users}

    location_tree = get_location_tree_json()

    index = 0
    status = 'success'
    tasks_status_map[task_id]['status'] = 'running'

    excel_row_datas = []
    exceed_flag = False
    dtable_col_name_to_column = {col['name']: col for col in base_columns}
    for row in ws.rows:
        if index > BIG_DATA_ROW_UPDATE_LIMIT:
            exceed_flag = True
            break

        try:
            if index > 0:  # skip header row
                row_list = [r.value for r in row]
                row_data = dict(zip(excel_columns, row_list))
                pop_col_lists = []
                for col_name, value in row_data.items():
                    col_type = column_name_type_map.get(col_name)
                    col_data = column_name_data_map.get(col_name)
                    if not col_type or col_type in AUTO_GENERATED_COLUMNS:
                        pop_col_lists.append(col_name)
                        continue
                    elif col_type == ColumnTypes.DATE and col_data:
                        row_data[col_name] = format_date(str(value), col_data.get('format'))

                row_data1 = deepcopy(row_data)
                for col_name in pop_col_lists:
                    row_data1.pop(col_name, None)

                excel_row_datas.append(row_data1)
                if len(excel_row_datas) >= 100:
                    rows_for_import, rows_for_update, excel_select_column_options = handle_excel_row_datas(
                        db_handler, table_name,
                        excel_row_datas, ref_columns,
                        column_name_type_map, column_name_data_map, name_to_email, location_tree,
                        is_insert_new_data
                    )
                    is_added_options = add_column_options(base, table_name, excel_select_column_options, dtable_col_name_to_column)
                    if is_added_options:
                        # if column option is added, update base_columns
                        base_columns = base.list_columns(table_name)
                        dtable_col_name_to_column = {col['name']: col for col in base_columns}

                    if is_insert_new_data and rows_for_import:
                        db_handler.insert_rows(table_name, rows_for_import)
                    if rows_for_update:
                        db_handler.batch_update_rows(table_name, rows_for_update)
                    excel_row_datas = []
                tasks_status_map[task_id]['rows_handled'] = total_count
                total_count += 1
            index += 1
        except Exception as err:
            tasks_status_map[task_id]['err_msg'] = 'Row updated error'
            tasks_status_map[task_id]['status'] = 'terminated'
            tasks_status_map[task_id]['err_code'] = ROW_INSERT_ERROR_CODE
            dtable_io_logger.exception(err)
            os.remove(file_path)
            return

    if excel_row_datas:
        rows_for_import, rows_for_update, excel_select_column_options = handle_excel_row_datas(
            db_handler, table_name,
            excel_row_datas, ref_columns,
            column_name_type_map, column_name_data_map, name_to_email, location_tree,
            is_insert_new_data
        )
        add_column_options(base, table_name, excel_select_column_options, dtable_col_name_to_column)

        if is_insert_new_data and rows_for_import:
            db_handler.insert_rows(table_name, rows_for_import)
        if rows_for_update:
            db_handler.batch_update_rows(table_name, rows_for_update)
        total_count += len(excel_row_datas)

    if exceed_flag:
        tasks_status_map[task_id]['err_msg'] = 'Number of rows exceeds %s limit' % BIG_DATA_ROW_UPDATE_LIMIT
        tasks_status_map[task_id]['status'] = 'terminated'
        tasks_status_map[task_id]['err_code'] = ROW_EXCEED_ERROR_CODE
        tasks_status_map[task_id]['rows_imported'] = total_count
        os.remove(file_path)
        return

    tasks_status_map[task_id]['status'] = status
    tasks_status_map[task_id]['rows_handled'] = total_count
    os.remove(file_path)
    return


def export_big_data_to_excel(dtable_uuid, table_id, view_id, username, name, task_id, tasks_status_map, repo_id, is_support_image=False):
    from dtable_events.dtable_io import dtable_io_logger

    # init task_status_map for exporting big data process
    tasks_status_map[task_id] = {
        'status': 'initializing',
        'err_msg': '',
        'handled_row_count': 0,
        'total_row_count': 0,
    }

    target_dir = TEMP_EXPORT_VIEW_DIR + dtable_uuid
    if not os.path.isdir(target_dir):
        os.makedirs(target_dir)

    try:
        nicknames = get_related_nicknames_from_dtable(dtable_uuid)
    except Exception as e:
        dtable_io_logger.error('get nicknames. ERROR: {}'.format(e))
        return
    email2nickname = {nickname['email']: nickname['name'] for nickname in nicknames}

    try:
        metadata = get_metadata_from_dtable_server(dtable_uuid, username)
    except Exception as e:
        dtable_io_logger.error('get metadata. ERROR: {}'.format(e))
        return

    target_table = {}
    target_view = {}
    for table in metadata.get('tables', []):
        if table.get('_id', '') == table_id:
            target_table = table
            break

    if not target_table:
        dtable_io_logger.warning('Table %s not found.' % table_id)
        return

    for view in target_table.get('views', []):
        if view.get('_id', '') == view_id:
            target_view = view
            break
    if not target_view:
        dtable_io_logger.warning('View %s not found.' % view_id)
        return

    table_name = target_table.get('name', '')
    view_name = target_view.get('name', '')
    cols = target_table.get('columns', [])
    hidden_cols_key = target_view.get('hidden_columns', [])
    summary_configs = target_table.get('summary_configs', {})
    cols_without_hidden = []
    summary_col_info = {}
    row_height = target_view.get('row_height', 'default')
    header_height = 'default'
    header_settings = target_table.get('header_settings')
    if header_settings:
        header_height = header_settings.get('header_height', 'default')
    for col in cols:
        if col.get('key', '') not in hidden_cols_key:
            cols_without_hidden.append(col)
        if summary_configs.get(col.get('key')):
            summary_col_info.update({col.get('name'): summary_configs.get(col.get('key'))})

    sheet_name = table_name + ('_' + view_name if view_name else '')
    sheet_name = escape_sheet_name(sheet_name)
    excel_name = name + '_' + table_name + ('_' + view_name if view_name else '') + '.xlsx'
    target_path = os.path.join(target_dir, excel_name)

    images_target_dir = os.path.join(IMAGE_TMP_DIR, dtable_uuid, str(uuid.uuid4()))
    image_param = {'num': 0, 'is_support': is_support_image, 'images_target_dir': images_target_dir}

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(sheet_name)

    dtable_db_api = DTableDBAPI(username, dtable_uuid, INNER_DTABLE_DB_URL)
    try:
        row_count_sql = 'select count(*) as total_count from `%s`' % table_name
        result, _ = dtable_db_api.query(row_count_sql, server_only=False)
        total_row_count = result[0].get('total_count', 0)
    except Exception as e:
        dtable_io_logger.error('get big data rows count error: %s', e)
        tasks_status_map[task_id]['status'] = 'terminated'
        tasks_status_map[task_id]['err_msg'] = 'get big data rows count failed'
        return

    column_name_to_column = {col.get('name'): col for col in cols}

    # exported row number should less than ARCHIVE_VIEW_EXPORT_ROW_LIMIT
    if total_row_count > int(ARCHIVE_VIEW_EXPORT_ROW_LIMIT):
        total_row_count = int(ARCHIVE_VIEW_EXPORT_ROW_LIMIT)

    tasks_status_map[task_id]['total_row_count'] = total_row_count

    filter_conditions = {}
    filter_conditions['sorts'] = target_view.get('sorts')
    filter_conditions['filters'] = target_view.get('filters')
    filter_conditions['filter_conjunction'] = target_view.get('filter_conjunction')

    offset = 10000
    start = 0
    while True:
        # exported row number should less than ARCHIVE_VIEW_EXPORT_ROW_LIMIT
        if (start + offset) > total_row_count:
            offset = total_row_count - start

        filter_conditions['start'] = start
        filter_conditions['limit'] = offset
        sql = filter2sql(table_name, cols, filter_conditions, by_group=False)
        response_rows, db_metadata = dtable_db_api.query(sql, convert=True, server_only=False)

        row_num = start
        try:
            write_xls_with_type(response_rows, email2nickname, ws, row_num, dtable_uuid, repo_id, image_param, cols_without_hidden, column_name_to_column, row_height=row_height, header_height=header_height, is_big_data_view=True)
        except Exception as e:
            dtable_io_logger.exception(e)
            dtable_io_logger.error('head_list = {}\n{}'.format(cols_without_hidden, e))
            tasks_status_map[task_id]['status'] = 'terminated'
            tasks_status_map[task_id]['err_msg'] = 'write xls error'
            return

        start += offset
        tasks_status_map[task_id]['handled_row_count'] = start
        tasks_status_map[task_id]['status'] = 'running'

        if start >= total_row_count or len(response_rows) < offset:
            break

    tasks_status_map[task_id]['status'] = 'success'
    wb.save(target_path)
    # remove tmp images
    try:
        shutil.rmtree(images_target_dir)
    except:
        pass


def export_app_table_page_to_excel(dtable_uuid, repo_id, table_id, username, app_name, page_name, filter_condition_groups, shown_column_keys, task_id, tasks_status_map, is_support_image=False):
    from dtable_events.dtable_io import dtable_io_logger

    tasks_status_map[task_id] = {
        'status': 'initializing',
        'err_msg': '',
        'handled_row_count': 0,
        'total_row_count': 0,
    }

    target_dir = TEMP_EXPORT_VIEW_DIR + dtable_uuid
    if not os.path.isdir(target_dir):
        os.makedirs(target_dir)

    try:
        nicknames = get_related_nicknames_from_dtable(dtable_uuid)
    except Exception as e:
        dtable_io_logger.error('get nicknames. ERROR: {}'.format(e))
        return
    email2nickname = {nickname['email']: nickname['name'] for nickname in nicknames}

    try:
        metadata = get_metadata_from_dtable_server(dtable_uuid, username)
    except Exception as e:
        dtable_io_logger.error('get metadata. ERROR: {}'.format(e))
        return

    target_table = {}
    for table in metadata.get('tables', []):
        if table.get('_id', '') == table_id:
            target_table = table
            break

    if not target_table:
        dtable_io_logger.warning('Table %s not found.' % table_id)
        return

    cols = target_table.get('columns', [])
    table_name = target_table.get('name', '')
    cols_without_hidden = []
    column_name_to_column = {}
    for col in cols:
        column_name_to_column[col.get('name')] = col
        if col.get('key', '') in shown_column_keys:
            cols_without_hidden.append(col)

    sheet_name = escape_sheet_name(page_name)
    excel_name = app_name + '_' + page_name + '.xlsx'
    target_path = os.path.join(target_dir, excel_name)
    images_target_dir = os.path.join(IMAGE_TMP_DIR, dtable_uuid, str(uuid.uuid4()))
    image_param = {'num': 0, 'is_support': is_support_image, 'images_target_dir': images_target_dir}

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(sheet_name)

    dtable_db_api = DTableDBAPI(username, dtable_uuid, INNER_DTABLE_DB_URL)
    try:
        sql_gen = BaseSQLGenerator(table_name, cols, filter_condition_groups=filter_condition_groups)
        row_count_sql = f'select count(*) as total_count from `{table_name}` {sql_gen._groupfilter2sql()}'
        result, _ = dtable_db_api.query(row_count_sql, server_only=False)
        total_row_count = result[0].get('total_count', 0)
    except Exception as e:
        dtable_io_logger.error('get big data rows count error: %s', e)
        tasks_status_map[task_id]['status'] = 'terminated'
        tasks_status_map[task_id]['err_msg'] = 'get big data rows count failed'
        return

    # exported row number should less than APP_TABLE_EXPORT_EXCEL_ROW_LIMIT
    if total_row_count > int(APP_TABLE_EXPORT_EXCEL_ROW_LIMIT):
        total_row_count = int(APP_TABLE_EXPORT_EXCEL_ROW_LIMIT)

    tasks_status_map[task_id]['total_row_count'] = total_row_count

    limit = 10000
    start = 0
    while True:
        # exported row number should less than APP_TABLE_EXPORT_EXCEL_ROW_LIMIT
        if (start + limit) > total_row_count:
            limit = total_row_count - start

        
        filter_condition_groups.update({
            'limit': limit,
            'start': start
        })

        sql = filter2sql(table_name, cols, filter_condition_groups, by_group=True)
        
        response_rows, _ = dtable_db_api.query(sql, convert=True, server_only=False)
        
        row_num = start
        try:
            write_xls_with_type(response_rows, email2nickname, ws, row_num, dtable_uuid, repo_id, image_param, cols_without_hidden, column_name_to_column, is_big_data_view=True)
        except Exception as e:
            dtable_io_logger.exception(e)
            dtable_io_logger.error('head_list = {}\n{}'.format(cols_without_hidden, e))
            tasks_status_map[task_id]['status'] = 'terminated'
            tasks_status_map[task_id]['err_msg'] = 'write xls error'
            return

        start += limit
        tasks_status_map[task_id]['handled_row_count'] = start
        tasks_status_map[task_id]['status'] = 'running'

        if start >= total_row_count or len(response_rows) < limit:
            break

    tasks_status_map[task_id]['status'] = 'success'
    wb.save(target_path)
    # remove tmp images
    try:
        shutil.rmtree(images_target_dir)
    except:
        pass
