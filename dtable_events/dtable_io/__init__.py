import os
import shutil
import uuid
import io
import json

import requests
from datetime import datetime
from sqlalchemy import text

from seaserv import seafile_api

from dtable_events.app.config import DTABLE_WEB_SERVICE_URL, INNER_DTABLE_SERVER_URL
from dtable_events.dtable_io.big_data import import_excel_to_db, update_excel_to_db, export_big_data_to_excel, \
    export_app_table_page_to_excel
from dtable_events.dtable_io.utils import import_archive_from_src_dtable, post_big_data_screen_app_zip_file, \
    post_dtable_json, post_asset_files, \
    download_files_to_path, create_forms_from_src_dtable, copy_src_forms_to_json, prepare_asset_file_folder, \
    prepare_dtable_json_from_memory, update_page_design_static_image, \
    copy_src_auto_rules_to_json, create_auto_rules_from_src_dtable, sync_app_users_to_table, \
    copy_src_workflows_to_json, create_workflows_from_src_dtable, copy_src_external_app_to_json,\
    create_external_apps_from_src_dtable, zip_big_data_screen, post_big_data_screen_zip_file, \
    export_page_design_dir_to_path, update_page_design_content_to_path, upload_page_design, \
    download_page_design_file, zip_big_data_screen_app, prepare_asset_files_download, copy_src_archive_backup
from dtable_events.db import init_db_session_class
from dtable_events.dtable_io.excel import parse_excel_csv_to_json, import_excel_csv_by_dtable_server, \
    append_parsed_file_by_dtable_server, parse_append_excel_csv_upload_file_to_json, \
    import_excel_csv_add_table_by_dtable_server, update_parsed_file_by_dtable_server, \
    parse_update_excel_upload_excel_to_json, parse_update_csv_upload_csv_to_json, parse_and_import_excel_csv_to_dtable, \
    parse_and_import_excel_csv_to_table, parse_and_update_file_to_table, parse_and_append_excel_csv_to_table
from dtable_events.convert_page.utils import get_chrome_data_dir, get_driver, open_page_view, wait_page_view, \
    gen_page_design_pdf_view_url, gen_document_pdf_view_url
from dtable_events.statistics.db import save_email_sending_records, batch_save_email_sending_records
from dtable_events.data_sync.data_sync_utils import run_sync_emails
from dtable_events.utils import is_valid_email, uuid_str_to_36_chars, gen_file_upload_url, uuid_str_to_32_chars
from dtable_events.utils.dtable_server_api import DTableServerAPI, BaseExceedsException
from dtable_events.utils.dtable_web_api import DTableWebAPI
from dtable_events.utils.exception import ExcelFormatError
from dtable_events.utils.email_sender import toggle_send_email
from dtable_events.dtable_io.utils import clear_tmp_dir, clear_tmp_file, clear_tmp_files_and_dirs, \
    SDOC_IMAGES_DIR, get_seadoc_download_link, gen_seadoc_base_dir, export_sdoc_prepare_images_folder,\
    batch_upload_sdoc_images, get_documents_config, save_documents_config
from dtable_events.app.log import setup_logger

dtable_io_logger = setup_logger('dtable_events_io', propagate=False)
dtable_message_logger = setup_logger('dtable_events_message', propagate=False)
dtable_data_sync_logger = setup_logger('dtable_events_data_sync', propagate=False)
dtable_plugin_email_logger = setup_logger('dtable_events_plugin_email', propagate=False)


def add_task_id_to_log(log_msg, task_id=None):
    return f'task [{task_id}] - {log_msg}' if task_id else log_msg


def get_dtable_export_content(username, repo_id, workspace_id, dtable_uuid, asset_dir_id, ignore_archive_backup, config, task_id):
    """
    1. prepare file content at /tmp/dtable-io/<dtable_id>/dtable_asset/...
    2. make zip file
    3. return zip file's content
    """
    task_result = {
        'warnings': []
    }

    dtable_io_logger.info(add_task_id_to_log(f'Start prepare /tmp/dtable-io/{dtable_uuid}/zip_file.zip for export DTable.', task_id))

    tmp_file_path = os.path.join('/tmp/dtable-io', dtable_uuid,
                                 'dtable_asset/')  # used to store asset files and json from file_server
    tmp_zip_path = os.path.join('/tmp/dtable-io', dtable_uuid, 'zip_file') + '.zip'  # zip path of zipped xxx.dtable

    db_session = init_db_session_class(config)()

    dtable_io_logger.info(add_task_id_to_log('Clear tmp dirs and files before prepare.', task_id))
    clear_tmp_files_and_dirs(tmp_file_path, tmp_zip_path)
    os.makedirs(tmp_file_path, exist_ok=True)
    # import here to avoid circular dependency

    # 1. create 'content.json' from 'xxx.dtable'
    dtable_io_logger.info(add_task_id_to_log('Create content.json file.', task_id))
    try:
        prepare_dtable_json_from_memory(workspace_id, dtable_uuid, username, tmp_file_path)
    except Exception as e:
        error_msg = 'prepare dtable json failed. ERROR: {}'.format(e)
        dtable_io_logger.exception(add_task_id_to_log(error_msg, task_id))
        db_session.close()
        raise Exception(error_msg)

    # 2. get asset file folder, asset could be empty
    if asset_dir_id:
        dtable_io_logger.info(add_task_id_to_log('Create asset dir.', task_id))
        try:
            prepare_asset_file_folder(username, repo_id, dtable_uuid, asset_dir_id, tmp_file_path, task_id)
        except Exception as e:
            error_msg = 'dtable: {} create asset folder failed. ERROR: {}'.format(dtable_uuid, e)
            dtable_io_logger.exception(add_task_id_to_log(error_msg, task_id))
            task_result['warnings'].append({'error': error_msg})

    # 3. copy forms
    try:
        copy_src_forms_to_json(dtable_uuid, tmp_file_path, db_session)
    except Exception as e:
        error_msg = 'copy forms failed. ERROR: {}'.format(e)
        dtable_io_logger.exception(add_task_id_to_log(error_msg, task_id))

    # 4. copy automation rules
    try:
        copy_src_auto_rules_to_json(dtable_uuid, tmp_file_path, db_session)
    except Exception as e:
        error_msg = 'copy automation rules failed. ERROR: {}'.format(e)
        dtable_io_logger.exception(add_task_id_to_log(error_msg, task_id))

    # 5. copy workflows
    try:
        copy_src_workflows_to_json(dtable_uuid, tmp_file_path, db_session)
    except Exception as e:
        error_msg = 'copy workflows failed. ERROR: {}'.format(e)
        dtable_io_logger.exception(add_task_id_to_log(error_msg, task_id))

    # 5. copy external app
    try:
        copy_src_external_app_to_json(dtable_uuid, tmp_file_path, db_session)
    except Exception as e:
        error_msg = 'copy external apps failed. ERROR: {}'.format(e)
        dtable_io_logger.exception(add_task_id_to_log(error_msg, task_id))

    # 6. archive backup
    if not ignore_archive_backup:
        dtable_io_logger.info(add_task_id_to_log('Export archive backup', task_id))
        try:
            copy_src_archive_backup(dtable_uuid, tmp_file_path, task_id)
        except Exception as e:
            error_msg = 'export archive backup failed. ERROR: {}'.format(e)
            dtable_io_logger.error(add_task_id_to_log(error_msg, task_id))
            clear_tmp_files_and_dirs(tmp_file_path, tmp_zip_path)
            if db_session:
                db_session.close()

    """
    /tmp/dtable-io/<dtable_uuid>/dtable_asset/
                                    |- asset/
                                    |- content.json

    we zip /tmp/dtable-io/<dtable_uuid>/dtable_asset/ to /tmp/dtable-io/<dtable_id>/zip_file.zip and download it
    notice than make_archive will auto add .zip suffix to /tmp/dtable-io/<dtable_id>/zip_file
    """
    dtable_io_logger.info(add_task_id_to_log('Make zip file for download...', task_id))
    try:
        shutil.make_archive('/tmp/dtable-io/' + dtable_uuid + '/zip_file', "zip", root_dir=tmp_file_path)
    except Exception as e:
        error_msg = 'make zip failed. ERROR: {}'.format(e)
        dtable_io_logger.error(add_task_id_to_log(error_msg, task_id))
        db_session.close()
        raise Exception(error_msg)

    db_session.close()

    dtable_io_logger.info(add_task_id_to_log('Create /tmp/dtable-io/{}/zip_file.zip success!'.format(dtable_uuid), task_id))
    # we remove '/tmp/dtable-io/<dtable_uuid>' in dtable web api

    # if export dtable with archive backup, need to post .zip to seafile repo which dtable belongs to and to send a notificaition to user
    if not ignore_archive_backup and username != 'export_dtable_command':
        dtable_io_logger.info(add_task_id_to_log(f'Post dtable file to attachments.', task_id))
        files_dir_path = f'/asset/{uuid_str_to_36_chars(dtable_uuid)}/files/{str(datetime.today())[:7]}'
        if not seafile_api.get_dir_id_by_path(repo_id, files_dir_path):
            seafile_api.mkdir_with_parents(repo_id, '/', files_dir_path[1:], username)
        # query dtable name
        try:
            dtable_name = db_session.execute(text(f"SELECT name FROM dtables WHERE uuid='{uuid_str_to_32_chars(dtable_uuid)}'")).fetchone().name
        except Exception as e:
            dtable_io_logger.error(add_task_id_to_log(f'query dtable {dtable_uuid} name error {e}', task_id))
            clear_tmp_files_and_dirs(tmp_file_path, tmp_zip_path)
            if db_session:
                db_session.close()
            return
        # post file to seafile
        dtable_file_name = f"{dtable_name} - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.dtable"
        try:
            seafile_api.post_file(repo_id, tmp_zip_path, files_dir_path, dtable_file_name, username)
        except Exception as e:
            dtable_io_logger.error(add_task_id_to_log(f'post dtable file to attachments error {e}', task_id))
            clear_tmp_files_and_dirs(tmp_file_path, tmp_zip_path)
            if db_session:
                db_session.close()
            return
        # send notification
        dtable_web_api = DTableWebAPI(DTABLE_WEB_SERVICE_URL)
        detail = {
            'dtable_uuid': uuid_str_to_32_chars(dtable_uuid),
            'dtable_name': dtable_name,
            'file_name': dtable_file_name,
            'file_path': os.path.join(files_dir_path[len(f'/asset/{uuid_str_to_36_chars(dtable_uuid)}'):], dtable_file_name)
        }
        try:
            dtable_web_api.internal_add_notification([username], 'export_dtable_completed', detail)
        except Exception as e:
            dtable_io_logger.error(add_task_id_to_log(f"send notification to {username} error: {e}", task_id))
            if db_session:
                db_session.close()
        clear_tmp_files_and_dirs(tmp_file_path, tmp_zip_path)

    return task_result


def get_dtable_export_content_folder(username, repo_id, workspace_id, dtable_uuid, asset_dir_id, ignore_archive_backup, config, folder_path, task_id):
    """
    like `get_dtable_export_content` but to export to `folder_path` and not to archive
    """
    task_result = {
        'warnings': []
    }

    db_session = init_db_session_class(config)()

    # if not os.path.isdir(folder_path):
    if not os.path.exists(folder_path):
        dtable_io_logger.info(add_task_id_to_log(f'Start prepare {folder_path} for export DTable.', task_id))
        os.makedirs(folder_path, exist_ok=True)
    else:
        if os.path.isdir(folder_path):
            dtable_io_logger.info(add_task_id_to_log(f'Path {folder_path} exists for export DTable.', task_id))
        else:
            dtable_io_logger.error(add_task_id_to_log(f'Path {folder_path} exists but is not a dir', task_id))
            return

    # import here to avoid circular dependency

    # 1. create 'content.json' from 'xxx.dtable'
    dtable_io_logger.info(add_task_id_to_log('Create content.json file.', task_id))
    try:
        prepare_dtable_json_from_memory(workspace_id, dtable_uuid, username, folder_path)
    except Exception as e:
        error_msg = 'prepare dtable json failed. ERROR: {}'.format(e)
        dtable_io_logger.exception(add_task_id_to_log(error_msg, task_id))
        db_session.close()
        raise Exception(error_msg)

    # 2. get asset file folder, asset could be empty
    if asset_dir_id:
        dtable_io_logger.info(add_task_id_to_log('Create asset dir.', task_id))
        try:
            prepare_asset_files_download(username, repo_id, dtable_uuid, asset_dir_id, folder_path, task_id)
        except Exception as e:
            error_msg = 'dtable: {} create asset folder failed. ERROR: {}'.format(dtable_uuid, e)
            dtable_io_logger.exception(add_task_id_to_log(error_msg, task_id))
            task_result['warnings'].append({'error': error_msg})

    # 3. copy forms
    try:
        copy_src_forms_to_json(dtable_uuid, folder_path, db_session)
    except Exception as e:
        error_msg = 'copy forms failed. ERROR: {}'.format(e)
        dtable_io_logger.error(add_task_id_to_log(error_msg, task_id))

    # 4. copy automation rules
    try:
        copy_src_auto_rules_to_json(dtable_uuid, folder_path, db_session)
    except Exception as e:
        error_msg = 'copy automation rules failed. ERROR: {}'.format(e)
        dtable_io_logger.error(add_task_id_to_log(error_msg, task_id))

    # 5. copy workflows
    try:
        copy_src_workflows_to_json(dtable_uuid, folder_path, db_session)
    except Exception as e:
        error_msg = 'copy workflows failed. ERROR: {}'.format(e)
        dtable_io_logger.error(add_task_id_to_log(error_msg, task_id))

    # 5. copy external app
    try:
        copy_src_external_app_to_json(dtable_uuid, folder_path, db_session)
    except Exception as e:
        error_msg = 'copy external apps failed. ERROR: {}'.format(e)
        dtable_io_logger.error(add_task_id_to_log(error_msg, task_id))

    db_session.close()

    # 6. archive backup
    if not ignore_archive_backup:
        dtable_io_logger.info(add_task_id_to_log('Export archive backup', task_id))
        try:
            copy_src_archive_backup(dtable_uuid, folder_path, task_id)
        except Exception as e:
            error_msg = 'export archive backup failed. ERROR: {}'.format(e)
            dtable_io_logger.error(add_task_id_to_log(error_msg, task_id))
            clear_tmp_files_and_dirs(folder_path, folder_path)
            if db_session:
                db_session.close()

    return task_result


def post_dtable_import_files(username, repo_id, workspace_id, dtable_uuid, dtable_file_name, in_storage,
                             can_use_automation_rules, can_use_workflows, can_use_external_apps, can_import_archive, owner, org_id, config, task_id):
    """
    post files at /tmp/<dtable_uuid>/dtable_zip_extracted/ to file server
    unzip django uploaded tmp file is suppose to be done in dtable-web api.
    """
    dtable_io_logger.info(add_task_id_to_log(f'Start import DTable: {dtable_uuid}.', task_id))

    extracted_path = os.path.join('/tmp/dtable-io', dtable_uuid, 'dtable_zip_extracted')

    db_session = init_db_session_class(config)()

    dtable_io_logger.info(add_task_id_to_log('Prepare dtable json file and post it at file server.', task_id))
    try:
        dtable_content = post_dtable_json(username, repo_id, workspace_id, dtable_uuid, dtable_file_name, in_storage, extracted_path, db_session)
    except Exception as e:
        dtable_io_logger.exception(add_task_id_to_log(f'post dtable json failed. ERROR: {e}', task_id))
        db_session.close()
        raise e

    dtable_io_logger.info(add_task_id_to_log('Post asset files in tmp path to file server.', task_id))
    try:
        post_asset_files(repo_id, dtable_uuid, username, extracted_path)
    except Exception as e:
        dtable_io_logger.exception(add_task_id_to_log(f'post asset files failed. ERROR: {e}', task_id))

    dtable_io_logger.info(add_task_id_to_log('create forms from src dtable.', task_id))
    try:
        create_forms_from_src_dtable(workspace_id, dtable_uuid, extracted_path, db_session)
    except Exception as e:
        dtable_io_logger.exception(add_task_id_to_log(f'create forms failed. ERROR: {e}', task_id))

    old_new_workflow_token_dict = {}  # old new workflow token dict
    if can_use_workflows:
        dtable_io_logger.info(add_task_id_to_log('create workflows from src dtable.', task_id))
        try:
            create_workflows_from_src_dtable(username, workspace_id, repo_id, dtable_uuid, owner, org_id, old_new_workflow_token_dict, extracted_path, db_session)
        except Exception as e:
            dtable_io_logger.exception(add_task_id_to_log(f'create workflows failed. ERROR: {e}', task_id))

    if can_use_automation_rules:
        dtable_io_logger.info(add_task_id_to_log('create auto rules from src dtable.', task_id))
        try:
            create_auto_rules_from_src_dtable(username, workspace_id, repo_id, owner, org_id, dtable_uuid, old_new_workflow_token_dict, extracted_path, db_session)
        except Exception as e:
            dtable_io_logger.exception(add_task_id_to_log(f'create auto rules failed. ERROR: {e}', task_id))

    if can_use_external_apps:
        dtable_io_logger.info(add_task_id_to_log('create external apps from src dtable.', task_id))
        try:
            create_external_apps_from_src_dtable(username, dtable_uuid, db_session, org_id, workspace_id, repo_id, extracted_path)
        except Exception as e:
            dtable_io_logger.exception(add_task_id_to_log(f'create external apps failed. ERROR: {e}', task_id))

    archive_file_path = os.path.join(extracted_path, 'archive')
    include_archive = can_import_archive and os.path.exists(archive_file_path)
    if include_archive:
        dtable_io_logger.info(add_task_id_to_log('import archive backup from src dtable.', task_id))
        try:
            status = import_archive_from_src_dtable(username, dtable_uuid, extracted_path)
        except Exception as e:
            dtable_io_logger.exception(f"import archive for {dtable_uuid} error: {e}")
        else:
            if status == 'failure':
                dtable_io_logger.error(f"import archive for {dtable_uuid} status: {status}")
            else:
                dtable_io_logger.info(f"import archive for {dtable_uuid} status: {status}")

    try:
        if dtable_content:
            plugin_settings = dtable_content.get('plugin_settings', {})
            page_design_settings = plugin_settings.get('page-design', [])
            page_design_content_json_tmp_path = os.path.join(extracted_path, 'page-design')
            # handle different url in settings.py
            update_page_design_static_image(page_design_settings, repo_id, workspace_id, dtable_uuid, page_design_content_json_tmp_path, username)
    except Exception as e:
        dtable_io_logger.exception(add_task_id_to_log(f'update page design static image failed. ERROR: {e}', task_id))

    # remove extracted tmp file
    dtable_io_logger.info(add_task_id_to_log('Remove extracted tmp file.', task_id))
    try:
        shutil.rmtree(os.path.join('/tmp/dtable-io', dtable_uuid))
    except Exception as e:
        dtable_io_logger.error(add_task_id_to_log(f'rm extracted tmp file failed. ERROR: {e}', task_id))

    db_session.close()

    # send notifications if include_archive
    if include_archive:
        dtable_web_api = DTableWebAPI(DTABLE_WEB_SERVICE_URL)
        detail = {
            'dtable_uuid': uuid_str_to_32_chars(dtable_uuid),
            'dtable_name': os.path.splitext(dtable_file_name)[0]
        }
        try:
            dtable_web_api.internal_add_notification([username], 'import_dtable_completed', detail)
        except Exception as e:
            dtable_io_logger.error(f"send notification to user {username} import_dtable_completed error {e}")

    dtable_io_logger.info(add_task_id_to_log(f'Import DTable: {dtable_uuid} success!', task_id))


def post_dtable_import_files_folder(username, repo_id, workspace_id, dtable_uuid, folder_path, in_storage,
                             can_use_automation_rules, can_use_workflows, can_use_external_apps, can_import_archive, owner, org_id, config, task_id):
    """
    like `post_dtable_import_files` but import from `folder_path` and not to remove folder_path
    """
    dtable_io_logger.info(add_task_id_to_log(f'Start import DTable: {dtable_uuid}.', task_id))

    dtable_file_name = os.path.basename(folder_path)
    db_session = init_db_session_class(config)()

    dtable_io_logger.info(add_task_id_to_log('Prepare dtable json file and post it at file server.', task_id))
    try:
        dtable_content = post_dtable_json(username, repo_id, workspace_id, dtable_uuid, dtable_file_name, in_storage, folder_path, db_session)
    except Exception as e:
        dtable_io_logger.exception(add_task_id_to_log(f'post dtable json failed. ERROR: {e}', task_id))
        db_session.close()
        raise e

    dtable_io_logger.info(add_task_id_to_log('Post asset files in tmp path to file server.', task_id))
    try:
        post_asset_files(repo_id, dtable_uuid, username, folder_path)
    except Exception as e:
        dtable_io_logger.exception(add_task_id_to_log(f'post asset files failed. ERROR: {e}', task_id))

    dtable_io_logger.info(add_task_id_to_log('create forms from src dtable.', task_id))
    try:
        create_forms_from_src_dtable(workspace_id, dtable_uuid, folder_path, db_session)
    except Exception as e:
        dtable_io_logger.exception(add_task_id_to_log(f'create forms failed. ERROR: {e}', task_id))

    old_new_workflow_token_dict = {}  # old new workflow token dict
    if can_use_workflows:
        dtable_io_logger.info(add_task_id_to_log('create workflows from src dtable.', task_id))
        try:
            create_workflows_from_src_dtable(username, workspace_id, repo_id, dtable_uuid, owner, org_id, old_new_workflow_token_dict, folder_path, db_session)
        except Exception as e:
            dtable_io_logger.exception(add_task_id_to_log(f'create workflows failed. ERROR: {e}', task_id))

    if can_use_automation_rules:
        dtable_io_logger.info(add_task_id_to_log('create auto rules from src dtable.', task_id))
        try:
            create_auto_rules_from_src_dtable(username, workspace_id, repo_id, owner, org_id, dtable_uuid, old_new_workflow_token_dict, folder_path, db_session)
        except Exception as e:
            dtable_io_logger.exception(add_task_id_to_log(f'create auto rules failed. ERROR: {e}', task_id))

    if can_use_external_apps:
        dtable_io_logger.info(add_task_id_to_log('create external apps from src dtable.', task_id))
        try:
            create_external_apps_from_src_dtable(username, dtable_uuid, db_session, org_id, workspace_id, repo_id, folder_path)
        except Exception as e:
            dtable_io_logger.exception(add_task_id_to_log(f'create external apps failed. ERROR: {e}', task_id))

    archive_file_path = os.path.join(folder_path, 'archive')
    include_archive = can_import_archive and os.path.exists(archive_file_path)
    if can_import_archive and include_archive:
        dtable_io_logger.info(add_task_id_to_log('import archive backup from src dtable.', task_id))
        try:
            status = import_archive_from_src_dtable(username, dtable_uuid, folder_path)
        except Exception as e:
            dtable_io_logger.exception(f"import archive for {dtable_uuid} error: {e}")
        else:
            if status == 'failure':
                dtable_io_logger.error(f"import archive for {dtable_uuid} status: {status}")
            else:
                dtable_io_logger.info(f"import archive for {dtable_uuid} status: {status}")

    try:
        if dtable_content:
            plugin_settings = dtable_content.get('plugin_settings', {})
            page_design_settings = plugin_settings.get('page-design', [])
            page_design_content_json_tmp_path = os.path.join(folder_path, 'page-design')
            # handle different url in settings.py
            update_page_design_static_image(page_design_settings, repo_id, workspace_id, dtable_uuid, page_design_content_json_tmp_path, username)
    except Exception as e:
        dtable_io_logger.exception(add_task_id_to_log(f'update page design static image failed. ERROR: {e}', task_id))

    db_session.close()

    dtable_io_logger.info(add_task_id_to_log(f'Import DTable: {dtable_uuid} success!', task_id))


def get_dtable_export_asset_files(username, repo_id, dtable_uuid, files, task_id, config, files_map=None):
    """
    export asset files from dtable
    """
    handled_files = []
    for file in files:
        files_map_value = files_map.get(file) if files_map else None
        if files_map_value:
            files_map.pop(file)
        file = file.strip().strip('/')
        file = file[:file.find('?')] if '?' in file else file
        handled_files.append(file)
        if files_map_value:
            files_map[file] = files_map_value
    files = handled_files
    tmp_file_path = os.path.join('/tmp/dtable-io', dtable_uuid, 'asset-files',
                                 str(task_id))           # used to store files
    tmp_zip_path  = os.path.join('/tmp/dtable-io', dtable_uuid, 'asset-files',
                                 str(task_id)) + '.zip'  # zip those files

    clear_tmp_files_and_dirs(tmp_file_path, tmp_zip_path)
    os.makedirs(tmp_file_path, exist_ok=True)

    db_session = init_db_session_class(config)()
    try:
        # 1. download files to tmp_file_path
        download_files_to_path(username, repo_id, dtable_uuid, files, tmp_file_path, db_session, files_map)
        # 2. zip those files to tmp_zip_path
        shutil.make_archive(tmp_zip_path.split('.')[0], 'zip', root_dir=tmp_file_path)
    except Exception as e:
        dtable_io_logger.error('export asset files from dtable failed. ERROR: {}'.format(e))
    else:
        dtable_io_logger.info('export files from dtable: %s success!', dtable_uuid)
    finally:
        db_session.close()

def get_dtable_export_big_data_screen(username, repo_id, dtable_uuid, page_id, task_id):
    """
    parse json file in big data screen, and zip it for download
    """
    tmp_file_path = os.path.join('/tmp/dtable-io', dtable_uuid, 'big-data-screen', str(task_id))
    tmp_zip_path = os.path.join('/tmp/dtable-io', dtable_uuid, 'big-data-screen', str(task_id) + '.zip')
    clear_tmp_files_and_dirs(tmp_file_path, tmp_zip_path)
    os.makedirs(tmp_file_path.rstrip('/') + '/images', exist_ok=True)

    try:
        zip_big_data_screen(username, repo_id, dtable_uuid, page_id, tmp_file_path)
        shutil.make_archive(tmp_zip_path.split('.')[0], 'zip', root_dir=tmp_file_path)
    except Exception as e:
        dtable_io_logger.error('export big data screen from dtable failed. ERROR: {}'.format(e))
    else:
        dtable_io_logger.info('export big data screen from dtable: %s success!', dtable_uuid)

def import_big_data_screen(username, repo_id, dtable_uuid, page_id):
    """
    parse the zip in tmp folders and upload it
    """
    tmp_extracted_path = os.path.join('/tmp/dtable-io', dtable_uuid, 'big_data_screen_zip_extracted/')
    try:
        post_big_data_screen_zip_file(username, repo_id, dtable_uuid, page_id, tmp_extracted_path)
    except Exception as e:
        dtable_io_logger.error('import big data screen from dtable failed. ERROR: {}'.format(e))
    else:
        dtable_io_logger.info('import big data screen to dtable: %s success!', dtable_uuid)
    try:
        shutil.rmtree(tmp_extracted_path)
    except Exception as e:
        dtable_io_logger.error('rm extracted tmp file failed. ERROR: {}'.format(e))


def get_dtable_export_big_data_screen_app(username, repo_id, dtable_uuid, app_uuid, app_id, task_id, config):
    """
    parse json file in big data screen, and zip it for download
    """
    tmp_file_path = os.path.join('/tmp/dtable-io', dtable_uuid, 'big-data-screen', str(task_id))
    tmp_zip_path = os.path.join('/tmp/dtable-io', dtable_uuid, 'big-data-screen', str(task_id) + '.zip')
    clear_tmp_files_and_dirs(tmp_file_path, tmp_zip_path)
    os.makedirs(tmp_file_path.rstrip('/') + '/images', exist_ok=True)

    db_session = init_db_session_class(config)()

    try:
        zip_big_data_screen_app(username, repo_id, dtable_uuid, app_uuid, app_id, tmp_file_path, db_session)
        shutil.make_archive(tmp_zip_path.split('.')[0], 'zip', root_dir=tmp_file_path)
    except Exception as e:
        dtable_io_logger.exception('export big data screen from dtable failed. ERROR: {}'.format(e))
    else:
        dtable_io_logger.info('export big data screen from dtable: %s success!', dtable_uuid)
    finally:
        db_session.close()


def import_big_data_screen_app(username, repo_id, dtable_uuid, app_uuid, app_id, config):
    """
    parse the zip in tmp folders and upload it
    """
    tmp_extracted_path = os.path.join('/tmp/dtable-io', dtable_uuid, 'big_data_screen_zip_extracted', app_uuid)
    db_session = init_db_session_class(config)()
    try:
        post_big_data_screen_app_zip_file(username, repo_id, dtable_uuid, app_uuid, app_id, tmp_extracted_path, db_session)
    except Exception as e:
        dtable_io_logger.exception('import big data screen from dtable failed. ERROR: {}'.format(e))
    else:
        dtable_io_logger.info('import big data screen to dtable: %s success!', dtable_uuid)
    finally:
        db_session.close()
    try:
        shutil.rmtree(tmp_extracted_path)
    except Exception as e:
        dtable_io_logger.error('rm extracted tmp file failed. ERROR: {}'.format(e))

def parse_excel_csv(username, repo_id, file_name, file_type, parse_type, dtable_uuid, config):
    """
    parse excel or csv to json file, then upload json file to file server
    """
    dtable_io_logger.info('Start parse excel or csv: %s.%s.' % (file_name, file_type))
    try:
        parse_excel_csv_to_json(username, repo_id, file_name, file_type, parse_type, dtable_uuid)
    except ExcelFormatError as e:
        raise Exception(e)
    except Exception as e:
        dtable_io_logger.exception('parse excel or csv failed. ERROR: {}'.format(e))
    else:
        dtable_io_logger.info('parse excel %s.xlsx success!' % file_name)

def import_excel_csv(username, repo_id, dtable_uuid, dtable_name, included_tables, lang, config):
    """
    upload excel or csv json file to dtable-server
    """
    dtable_io_logger.info('Start import excel or csv: {}.'.format(dtable_uuid))
    try:
        import_excel_csv_by_dtable_server(username, repo_id, dtable_uuid, dtable_name, included_tables, lang)
    except BaseExceedsException as e:
        raise Exception(e.error_msg)
    except Exception as e:
        dtable_io_logger.error('import excel or csv failed. ERROR: {}'.format(e))
    else:
        dtable_io_logger.info('import excel or csv %s success!' % dtable_name)

def import_excel_csv_add_table(username, dtable_uuid, dtable_name, included_tables, lang, config):
    """
    add table, upload excel or csv json file to dtable-server
    """
    dtable_io_logger.info('Start import excel or csv add table: {}.'.format(dtable_uuid))
    try:
        import_excel_csv_add_table_by_dtable_server(username, dtable_uuid, dtable_name, included_tables, lang)
    except BaseExceedsException as e:
        raise Exception(e.error_msg)

    except Exception as e:
        dtable_io_logger.error('import excel or csv add table failed. dtable_uuid: %s, dtable_name: %s ERROR: %s' % (dtable_uuid, dtable_name, e))
        raise Exception('Import excel or csv error')
    else:
        dtable_io_logger.info('import excel or csv %s add table success!' % dtable_name)


def append_excel_csv_append_parsed_file(username, dtable_uuid, file_name, table_name):
    """
    upload excel or csv json file to dtable-server
    """
    dtable_io_logger.info('Start import excel or csv: {}.'.format(dtable_uuid))
    try:
        append_parsed_file_by_dtable_server(username, dtable_uuid, file_name, table_name)
    except Exception as e:
        dtable_io_logger.exception('append excel or csv failed. dtable_uuid: %s, table_name: %s ERROR:  %s' % (dtable_uuid, table_name, e))
        raise Exception('Import excel or csv error')
    else:
        dtable_io_logger.info('append excel or csv %s success!' % file_name)

def append_excel_csv_upload_file(username, file_name, dtable_uuid, table_name, file_type):
    """
    parse excel or csv to json file, then upload json file to file server
    """
    dtable_io_logger.info('Start parse append excel or csv: %s.%s' % (file_name, file_type))
    try:
        parse_append_excel_csv_upload_file_to_json(file_name, username, dtable_uuid, table_name, file_type)
    except ExcelFormatError:
        raise Exception('Excel format error')
    except Exception as e:
        dtable_io_logger.exception('parse append excel or csv failed. ERROR: {}'.format(e))
    else:
        dtable_io_logger.info('parse append excel or csv %s.%s success!' % (file_name, file_type))

def update_excel_csv_update_parsed_file(username, dtable_uuid, file_name, table_name, selected_columns, can_add_row, can_update_row):
    """
    upload excel/csv json file to dtable-server
    """
    dtable_io_logger.info('Start import file: {}.'.format(dtable_uuid))
    try:
        update_parsed_file_by_dtable_server(username, dtable_uuid, file_name, table_name, selected_columns, can_add_row, can_update_row)
    except Exception as e:
        dtable_io_logger.exception('update excel,csv failed. dtable_uuid: %s, table_name: %s ERROR: %s' % (dtable_uuid, table_name, e))
        raise Exception('Update excel or csv error')
    else:
        dtable_io_logger.info('update excel,csv %s success!' % file_name)

def update_excel_upload_excel(username, file_name, dtable_uuid, table_name):
    """
    parse excel to json file, then upload json file to file server
    """
    dtable_io_logger.info('Start parse update excel: %s.xlsx.' % file_name)
    try:
        parse_update_excel_upload_excel_to_json(file_name, username, dtable_uuid, table_name)
    except ExcelFormatError:
        raise Exception('Excel format error')
    except Exception as e:
        dtable_io_logger.exception('parse update excel failed. dtable_uuid: %s, table_name: %s ERROR: %s' % (dtable_uuid, table_name, e))
        raise Exception('Update excel or csv error')
    else:
        dtable_io_logger.info('parse update excel %s.xlsx success!' % file_name)

def update_csv_upload_csv(username, file_name, dtable_uuid, table_name):
    """
    parse csv to json file, then upload json file to file server
    """
    dtable_io_logger.info('Start parse update csv: %s.csv.' % file_name)
    try:
        parse_update_csv_upload_csv_to_json(file_name, username, dtable_uuid, table_name)
    except Exception as e:
        dtable_io_logger.exception('parse update csv failed. dtable_uuid: %s, table_name: %s ERROR: %s' % (dtable_uuid, table_name, e))
        raise Exception('Update excel or csv error')
    else:
        dtable_io_logger.info('parse update csv %s.csv success!' % file_name)


def import_excel_csv_to_dtable(username, repo_id, dtable_name, dtable_uuid, file_type, lang):
    """
    parse excel csv to json, then import excel csv to dtable
    """
    dtable_io_logger.info('Start import excel or csv: %s.%s to dtable.' % (dtable_name, file_type))
    try:
        parse_and_import_excel_csv_to_dtable(repo_id, dtable_name, dtable_uuid, username, file_type, lang)
    except BaseExceedsException as e:
        raise Exception(e.error_msg)
    except ExcelFormatError:
        raise Exception('Excel format error')
    except Exception as e:
        dtable_io_logger.exception('import excel or csv to dtable failed. dtable_uuid: %s, dtable_name: %s ERROR: %s' % (dtable_uuid, dtable_name, e))
        raise Exception('Import excel or csv error')
    else:
        dtable_io_logger.info('import excel or csv %s.%s to dtable success!' % (dtable_name, file_type))


def import_excel_csv_to_table(username, file_name, dtable_uuid, file_type, lang):
    """
    parse excel or csv to json, then import excel or csv to table
    """
    dtable_io_logger.info('Start import excel or csv: %s.%s to table.' % (file_name, file_type))
    try:
        parse_and_import_excel_csv_to_table(file_name, dtable_uuid, username, file_type, lang)
    except BaseExceedsException as e:
        raise Exception(e.error_msg)

    except ExcelFormatError:
        raise Exception('Excel format error')
    except Exception as e:
        dtable_io_logger.exception('import excel or csv to table failed.  dtable_uuid: %s, file_name: %s ERROR: %s' % (dtable_uuid, file_name, e))
        raise Exception('Import excel or csv error')
    else:
        dtable_io_logger.info('import excel or csv %s.%s to table success!' % (file_name, file_type))


def update_table_via_excel_csv(username, file_name, dtable_uuid, table_name, selected_columns, file_type, can_add_row, can_update_row):
    """
    update excel/csv file to table
    """
    dtable_io_logger.info('Start update file: %s.%s to table.' % (file_name, file_type))
    try:
        parse_and_update_file_to_table(file_name, username, dtable_uuid, table_name, selected_columns, file_type, can_add_row, can_update_row)
    except Exception as e:
        dtable_io_logger.exception('update file update to table failed. dtable_uuid: %s, table_name: %s ERROR: %s' % (dtable_uuid, table_name, e))
        raise Exception('Update excel or csv error')
    else:
        dtable_io_logger.info('update file %s.%s update to table success!' % (file_name, file_type))


def append_excel_csv_to_table(username, file_name, dtable_uuid, table_name, file_type):
    """
    parse excel or csv to json, then append excel or csv to table
    """
    dtable_io_logger.info('Start append excel or csv: %s.%s to table.' % (file_name, file_type))
    try:
        parse_and_append_excel_csv_to_table(username, file_name, dtable_uuid, table_name, file_type)
    except ExcelFormatError:
        raise Exception('Excel format error')
    except Exception as e:
        dtable_io_logger.exception('append excel or csv to table failed. dtable_uuid: %s, table_name: %s ERROR: %s' % (dtable_uuid, table_name, e))
        raise Exception('Import excel or csv error')
    else:
        dtable_io_logger.info('append excel or csv %s.%s to table success!' % (file_name, file_type))


def _get_upload_link_to_seafile(seafile_server_url, access_token, parent_dir="/"):
    upload_link_api_url = "%s%s" % (seafile_server_url.rstrip('/'),  '/api/v2.1/via-repo-token/upload-link/')
    headers = {
        'authorization': 'Token ' + access_token
    }
    params = {
        'path': parent_dir
    }
    response = requests.get(upload_link_api_url, headers=headers, params=params)
    return response.json()

def _upload_to_seafile(seafile_server_url, access_token, files, parent_dir="/", relative_path="", replace=None):
    upload_url = _get_upload_link_to_seafile(seafile_server_url, access_token, parent_dir)
    files_tuple_list = [('file', open(file, 'rb')) for file in files]
    files = files_tuple_list + [('parent_dir', parent_dir), ('relative_path', relative_path), ('replace', replace)]
    response = requests.post(upload_url, files=files)
    return response

def get_dtable_transfer_asset_files(username, repo_id, dtable_uuid, files, task_id, files_map, parent_dir, relative_path, replace, repo_api_token, seafile_server_url, config):
    tmp_file_path = os.path.join('/tmp/dtable-io/', dtable_uuid, 'transfer-files', str(task_id))
    os.makedirs(tmp_file_path, exist_ok=True)

    db_session = init_db_session_class(config)()
    try:
        # download files to local
        local_file_list = download_files_to_path(username, repo_id, dtable_uuid, files, tmp_file_path, db_session, files_map)
    except Exception as e:
        dtable_io_logger.error('export asset files from dtable failed. ERROR: {}'.format(e))
        if os.path.exists(tmp_file_path):
            shutil.rmtree(tmp_file_path)
        return
    finally:
        db_session.close()

    # upload files from local to seafile
    try:
        _upload_to_seafile(seafile_server_url, repo_api_token, local_file_list, parent_dir, relative_path, replace)
    except Exception as e:
        dtable_io_logger.error('transfer asset files from dtable failed. ERROR: {}'.format(e))

    # delete local files
    if os.path.exists(tmp_file_path):
        shutil.rmtree(tmp_file_path)

def send_wechat_msg(webhook_url, msg, msg_type="text"):
    if msg_type == "markdown":
        msg_format = {"msgtype": "markdown","markdown":{"content":msg}}
    else:
        msg_format = {"msgtype": "text", "text": {"content": msg}}
    result = {}
    try:
        requests.post(webhook_url, json=msg_format, headers={"Content-Type": "application/json"})
    except Exception as e:
        dtable_message_logger.error('Wechat sending failed. ERROR: {}'.format(e))
        result['err_msg'] = 'Webhook URL invalid'
    else:
        dtable_message_logger.info('Wechat sending success!')
    return result

def send_dingtalk_msg(webhook_url, msg, msg_type="text", msg_title=None):
    result = {}
    if msg_type == "markdown":
        if not msg_title:
            result['err_msg'] = 'msg_title invalid'
            dtable_message_logger.error('Dingtalk sending failed. ERROR: msg_title invalid')
            return result
        msg_format = {"msgtype": "markdown", "markdown": {"text": msg, "title": msg_title}}
    else:
        msg_format = {"msgtype": "text", "text": {"content": msg}}

    try:
        requests.post(webhook_url, json=msg_format, headers={"Content-Type": "application/json"})
    except Exception as e:
        dtable_message_logger.error('Dingtalk sending failed. ERROR: {}'.format(e))
        result['err_msg'] = 'Dingtalk URL invalid'
    else:
        dtable_message_logger.info('Dingtalk sending success!')
    return result

def send_notification_msg(emails, user_col_key, msg, dtable_uuid, username, table_id=None, row_id=None):
    result = {}
    try:
        dtable_server_api = DTableServerAPI(username, dtable_uuid, INNER_DTABLE_SERVER_URL)
        metadata = dtable_server_api.get_metadata()
        table = None
        for tmp_table in metadata['tables']:
            if tmp_table['_id'] == table_id:
                table = tmp_table
                break
        if not table:
            return

        target_row = dtable_server_api.get_row(table['name'], row_id)

        sending_list = emails
        if user_col_key:
            column = None
            for tmp_col in table['columns']:
                if tmp_col['key'] == user_col_key:
                    column = tmp_col
                    break

            user_col_info = column and target_row.get(column['name']) or None
            if user_col_info:
                if isinstance(user_col_info, list):
                    for user in user_col_info:
                        if user in sending_list:
                            continue
                        sending_list.append(user)
                else:
                    if user_col_info not in sending_list:
                        sending_list.append(user_col_info)

        detail = {
            'table_id': table_id or '',
            'msg': msg,
            'row_id_list': row_id and [row_id, ] or [],
        }
        user_msg_list = []
        for user in sending_list:
            if not is_valid_email(user):
                continue
            user_msg_list.append({
                'to_user': user,
                'msg_type': 'notification_rules',
                'detail': detail,
                })

        dtable_server_api.batch_send_notification(user_msg_list)
    except Exception as e:
        dtable_message_logger.error('Notification sending failed. ERROR: {}'.format(e))
        result['err_msg'] = 'Notification send failed'
    else:
        dtable_message_logger.info('Notification sending success!')
    return result

def convert_page_design_to_pdf(dtable_uuid, page_id, row_id, username, config):
    if not username:
        username = 'dtable-events'
    db_session = init_db_session_class(config)()
    sql = "SELECT `owner`, `org_id` FROM dtables d JOIN workspaces w ON d.workspace_id=w.id WHERE d.uuid=:dtable_uuid"
    try:
        result = db_session.execute(text(sql), {'dtable_uuid': uuid_str_to_32_chars(dtable_uuid)}).fetchone()
    except Exception as e:
        dtable_io_logger.error(f'query dtable {dtable_uuid} owner, org_id error {e}')
        return
    finally:
        db_session.close()
    kwargs = {'org_id': result.org_id, 'owner_id': result.owner}
    access_token = DTableServerAPI(username, dtable_uuid, INNER_DTABLE_SERVER_URL, kwargs=kwargs).internal_access_token
    target_dir = '/tmp/dtable-io/convert-page-to-pdf'
    if not os.path.isdir(target_dir):
        os.makedirs(target_dir)
    target_path = os.path.join(target_dir, '%s_%s_%s.pdf' % (dtable_uuid, page_id, row_id))

    chrome_data_dir_name = f'{dtable_uuid}-{page_id}-{row_id}'
    driver = get_driver(get_chrome_data_dir(chrome_data_dir_name))
    monitor_dom_id = 'page-design-render-complete'
    try:
        pdf_view_url = gen_page_design_pdf_view_url(dtable_uuid, page_id, access_token, row_id)

        session_id = open_page_view(driver, pdf_view_url)
        wait_page_view(driver, session_id, monitor_dom_id, row_id, target_path)
    except Exception as e:
        dtable_io_logger.exception('convert page_desgin, dtable: %s page: %s row: %s error: %s', dtable_uuid, page_id, row_id, e)
    finally:
        if os.path.exists(chrome_data_dir_name):
            shutil.rmtree(chrome_data_dir_name)
        driver.quit()


def convert_document_to_pdf(dtable_uuid, doc_uuid, row_id, username, config):
    if not username:
        username = 'dtable-events'
    db_session = init_db_session_class(config)()
    sql = "SELECT `owner`, `org_id` FROM dtables d JOIN workspaces w ON d.workspace_id=w.id WHERE d.uuid=:dtable_uuid"
    try:
        result = db_session.execute(text(sql), {'dtable_uuid': uuid_str_to_32_chars(dtable_uuid)}).fetchone()
    except Exception as e:
        dtable_io_logger.error(f'query dtable {dtable_uuid} owner, org_id error {e}')
        return
    finally:
        db_session.close()
    kwargs = {'org_id': result.org_id, 'owner_id': result.owner}
    access_token = DTableServerAPI(username, dtable_uuid, INNER_DTABLE_SERVER_URL, kwargs=kwargs).internal_access_token
    target_dir = '/tmp/dtable-io/convert-document-to-pdf'
    if not os.path.isdir(target_dir):
        os.makedirs(target_dir)
    target_path = os.path.join(target_dir, '%s_%s_%s.pdf' % (dtable_uuid, doc_uuid, row_id))

    chrome_data_dir_name = f'{dtable_uuid}-{doc_uuid}-{row_id}'
    driver = get_driver(get_chrome_data_dir(chrome_data_dir_name))
    monitor_dom_id = 'document-render-complete'
    try:
        pdf_view_url = gen_document_pdf_view_url(dtable_uuid, doc_uuid, access_token, row_id)
        session_id = open_page_view(driver, pdf_view_url)
        wait_page_view(driver, session_id, monitor_dom_id, row_id, target_path)
    except Exception as e:
        dtable_io_logger.exception('convert document, dtable: %s page: %s row: %s to pdf, error: %s', dtable_uuid, doc_uuid, row_id, e)
    finally:
        if os.path.exists(chrome_data_dir_name):
            shutil.rmtree(chrome_data_dir_name)
        driver.quit()


def convert_view_to_excel(dtable_uuid, table_id, view_id, username, id_in_org, user_department_ids_map, permission, name, repo_id, is_support_image=False):
    from dtable_events.dtable_io.utils import get_metadata_from_dtable_server, get_view_rows_from_dtable_server
    from dtable_events.dtable_io.excel import write_xls_with_type, TEMP_EXPORT_VIEW_DIR, IMAGE_TMP_DIR
    from dtable_events.dtable_io.utils import get_related_nicknames_from_dtable, escape_sheet_name
    import openpyxl

    target_dir = TEMP_EXPORT_VIEW_DIR + dtable_uuid
    if not os.path.isdir(target_dir):
        os.makedirs(target_dir)

    try:
        nicknames = get_related_nicknames_from_dtable(dtable_uuid)
    except Exception as e:
        dtable_io_logger.error('get nicknames. ERROR: {}'.format(e))
        return
    email2nickname = {nickname['email']: nickname['name'] for nickname in nicknames}

    try:
        metadata = get_metadata_from_dtable_server(dtable_uuid, username)
    except Exception as e:
        dtable_io_logger.error('get metadata. ERROR: {}'.format(e))
        return

    target_table = {}
    target_view = {}
    for table in metadata.get('tables', []):
        if table.get('_id', '') == table_id:
            target_table = table
            break

    if not target_table:
        dtable_io_logger.warning('Table %s not found.' % table_id)
        return

    for view in target_table.get('views', []):
        if view.get('_id', '') == view_id:
            target_view = view
            break
    if not target_view:
        dtable_io_logger.warning('View %s not found.' % view_id)
        return

    table_name = target_table.get('name', '')
    view_name = target_view.get('name', '')
    row_height = target_view.get('row_height', 'default')
    header_height = 'default'
    header_settings = target_table.get('header_settings')
    if header_settings:
        header_height = header_settings.get('header_height', 'default')

    cols = target_table.get('columns', [])
    hidden_cols_key = target_view.get('hidden_columns', [])
    summary_configs = target_table.get('summary_configs', {})
    cols_without_hidden = []
    summary_col_info = {}
    for col in cols:
        if col.get('key', '') not in hidden_cols_key:
            cols_without_hidden.append(col)
        if summary_configs.get(col.get('key')):
            summary_col_info.update({col.get('name'): summary_configs.get(col.get('key'))})

    images_target_dir = os.path.join(IMAGE_TMP_DIR, dtable_uuid, str(uuid.uuid4()))
    image_param = {'num': 0, 'is_support': is_support_image, 'images_target_dir': images_target_dir}

    sheet_name = table_name + ('_' + view_name if view_name else '')
    sheet_name = escape_sheet_name(sheet_name)
    excel_name = name + '_' + table_name + ('_' + view_name if view_name else '') + '.xlsx'

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(sheet_name)

    try:
        dtable_rows = get_view_rows_from_dtable_server(dtable_uuid, table_name, view_name, username, id_in_org, user_department_ids_map, permission)
    except BaseExceedsException as e:
        raise Exception(e.error_msg)

    column_name_to_column = {col.get('name'): col for col in cols}
    is_group_view = bool(target_view.get('groupbys'))

    params = (dtable_rows, email2nickname, ws, 0, dtable_uuid, repo_id, image_param, cols_without_hidden, column_name_to_column, is_group_view, summary_col_info, row_height, header_height)

    try:
        write_xls_with_type(*params)
    except Exception as e:
        dtable_io_logger.exception(e)
        dtable_io_logger.error('head_list = {}\n{}'.format(cols_without_hidden, e))
        return
    target_path = os.path.join(target_dir, excel_name)
    wb.save(target_path)
    # remove tmp images
    try:
        shutil.rmtree(images_target_dir)
    except:
        pass


def convert_table_to_excel(dtable_uuid, table_id, username, name, repo_id, is_support_image=False):
    from dtable_events.dtable_io.utils import get_metadata_from_dtable_server, get_rows_from_dtable_server
    from dtable_events.dtable_io.excel import write_xls_with_type, IMAGE_TMP_DIR
    from dtable_events.dtable_io.utils import get_related_nicknames_from_dtable, escape_sheet_name
    import openpyxl

    target_dir = '/tmp/dtable-io/export-table-to-excel/' + dtable_uuid
    if not os.path.isdir(target_dir):
        os.makedirs(target_dir)

    try:
        nicknames = get_related_nicknames_from_dtable(dtable_uuid)
    except Exception as e:
        dtable_io_logger.error('get nicknames. ERROR: {}'.format(e))
        return
    email2nickname = {nickname['email']: nickname['name'] for nickname in nicknames}

    try:
        metadata = get_metadata_from_dtable_server(dtable_uuid, username)
    except Exception as e:
        dtable_io_logger.error('get metadata. ERROR: {}'.format(e))
        return

    target_table = {}
    for table in metadata.get('tables', []):
        if table.get('_id', '') == table_id:
            target_table = table
            break

    if not target_table:
        dtable_io_logger.warning('Table %s not found.' % table_id)
        return

    table_name = target_table.get('name', '')
    cols = target_table.get('columns', [])
    header_height = 'default'
    header_settings = target_table.get('header_settings')
    if header_settings:
        header_height = header_settings.get('header_height', 'default')

    try:
        result_rows = get_rows_from_dtable_server(username, dtable_uuid, table_name)
    except BaseExceedsException as e:
        raise Exception(e.error_msg)
    column_name_to_column = {col.get('name'): col for col in cols}

    images_target_dir = os.path.join(IMAGE_TMP_DIR, dtable_uuid, str(uuid.uuid4()))
    image_param = {'num': 0, 'is_support': is_support_image, 'images_target_dir': images_target_dir}

    sheet_name = escape_sheet_name(table_name)
    excel_name = name + '_' + table_name + '.xlsx'
    target_path = os.path.join(target_dir, excel_name)
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(sheet_name)
    try:
        write_xls_with_type(result_rows, email2nickname, ws, 0, dtable_uuid, repo_id, image_param, cols, column_name_to_column, header_height=header_height)
    except Exception as e:
        dtable_io_logger.error('head_list = {}\n{}'.format(cols, e))
        return
    wb.save(target_path)
    # remove tmp images
    try:
        shutil.rmtree(images_target_dir)
    except:
        pass

def app_user_sync(dtable_uuid, app_name, app_id, table_name, table_id, username, config):
    dtable_io_logger.info('Start sync app %s users: to table %s.' % (app_name, table_name))
    db_session = init_db_session_class(config)()
    try:
        sync_app_users_to_table(dtable_uuid, app_id, table_name, table_id, username, db_session)
    except Exception as e:
        dtable_io_logger.exception('app user sync ERROR: {}'.format(e))
        raise Exception('app user sync ERROR: {}'.format(e))
    else:
        dtable_io_logger.info('app %s user sync success!' % app_name)
    finally:
        if db_session:
            db_session.close()


def email_sync(context, config):
    dtable_data_sync_logger.info('Start sync email to dtable %s, email table %s.' % (context.get('dtable_uuid'), context.get('detail',{}).get('email_table_id')))
    context['db_session_class'] = init_db_session_class(config)

    try:
        run_sync_emails(context)
    except Exception as e:
        dtable_data_sync_logger.exception('sync email ERROR: {}'.format(e))
    else:
        dtable_data_sync_logger.info('sync email success, sync_id: %s' % context.get('data_sync_id'))


def plugin_email_send_email(context, config=None):
    dtable_plugin_email_logger.info('Start send email by plugin %s, email table %s.' % (context.get('dtable_uuid'), context.get('table_info', {}).get('email_table_name')))

    dtable_uuid = context.get('dtable_uuid')
    username = context.get('username')
    repo_id = context.get('repo_id')
    workspace_id = context.get('workspace_id')

    table_info = context.get('table_info')
    email_info = context.get('email_info')
    account_id = context.get('account_id')

    thread_row_id = table_info.get('thread_row_id')
    email_row_id = table_info.get('email_row_id')
    email_table_name = table_info.get('email_table_name')
    thread_table_name = table_info.get('thread_table_name')

    # send email
    toggle_send_email(account_id, email_info, username, config)

    send_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    dtable_server_api = DTableServerAPI(username, dtable_uuid, INNER_DTABLE_SERVER_URL, server_url=DTABLE_WEB_SERVICE_URL,
                                        repo_id=repo_id, workspace_id=workspace_id)

    replied_email_row = dtable_server_api.get_row(email_table_name, email_row_id)

    thread_id = replied_email_row.get('Thread ID')

    html_message = email_info.get('html_message')
    if html_message:
        html_message = '```' + html_message + '```'

    email = {
        'cc': email_info.get('copy_to'),
        'From': email_info.get('from'),
        'Message ID': email_info.get('message_id'),
        'Reply to Message ID': email_info.get('in_reply_to'),
        'To': email_info.get('send_to'),
        'Subject': email_info.get('subject'),
        'Content': email_info.get('text_message'),
        'HTML Content': html_message,
        'Date': send_time,
        'Thread ID': thread_id,
    }

    # sync email table after sending
    metadata = dtable_server_api.get_metadata()

    tables = metadata.get('tables', [])
    email_table_id = ''
    link_table_id = ''
    for table in tables:
        if table.get('name') == email_table_name:
            email_table_id = table.get('_id')
        if table.get('name') == thread_table_name:
            link_table_id = table.get('_id')
        if email_table_id and link_table_id:
            break
    if not email_table_id or not link_table_id:
        dtable_plugin_email_logger.warning('Send email successfully but sync email table failure with email table: %s or link table: %s not found', email_table_name, thread_table_name)
        return

    email_link_id = dtable_server_api.get_column_link_id(email_table_name, 'Threads')

    email_row = dtable_server_api.append_row(email_table_name, email)
    email_row_id = email_row.get('_id')
    other_rows_ids = [thread_row_id]

    dtable_server_api.update_link(email_link_id, email_table_id, link_table_id, email_row_id, other_rows_ids)

    dtable_server_api.update_row(thread_table_name, thread_row_id, {'Last Updated': send_time})

def import_big_excel(username, dtable_uuid, table_name, file_path, task_id, tasks_status_map):
    """
    upload excel json file to dtable-db
    """

    dtable_io_logger.info('Start import big excel: {}.'.format(dtable_uuid))
    try:
        import_excel_to_db(username, dtable_uuid, table_name, file_path, task_id, tasks_status_map)
    except Exception as e:
        dtable_io_logger.exception('import big excel failed. ERROR: {}'.format(e))
    else:
        dtable_io_logger.info('import big excel %s.xlsx success!' % table_name)


def update_big_excel(username, dtable_uuid, table_name, file_path, ref_columns, is_insert_new_data, can_add_row, can_update_row, task_id, tasks_status_map):
    """
    upload excel json file to dtable-db
    """

    dtable_io_logger.info('Start update big excel: {}.'.format(dtable_uuid))
    try:
        update_excel_to_db(username, dtable_uuid, table_name, file_path, ref_columns, is_insert_new_data, task_id, tasks_status_map, can_add_row, can_update_row)
    except Exception as e:
        dtable_io_logger.exception('update big excel failed. ERROR: {}'.format(e))
    else:
        dtable_io_logger.info('update big excel %s.xlsx success!' % table_name)


def convert_big_data_view_to_excel(dtable_uuid, table_id, view_id, username, name, task_id, tasks_status_map, repo_id, is_support_image):
    dtable_io_logger.info('Start export big data view to excel: {}.'.format(dtable_uuid))
    try:
        export_big_data_to_excel(dtable_uuid, table_id, view_id, username, name, task_id, tasks_status_map, repo_id, is_support_image)
    except Exception as e:
        dtable_io_logger.error('export big data view failed. ERROR: {}'.format(e))
    else:
        dtable_io_logger.info('export big data table_id: %s, view_id: %s success!', table_id, view_id)


def export_page_design(repo_id, dtable_uuid, page_id, username):
    # prepare empty dir
    tmp_zip_dir = os.path.join('/tmp/dtable-io', 'page-design')
    os.makedirs(tmp_zip_dir, exist_ok=True)
    tmp_zip_path = os.path.join(tmp_zip_dir, f'{uuid_str_to_36_chars(dtable_uuid)}-{page_id}.zip')

    # download and save to path
    export_page_design_dir_to_path(repo_id, dtable_uuid, page_id, tmp_zip_path, username)


def import_page_design(repo_id, workspace_id, dtable_uuid, page_id, is_dir, username):
    # check file exists
    need_check_static = False
    try:
        if is_dir:
            tmp_page_path = os.path.join('/tmp/dtable-io', 'page-design', f'{uuid_str_to_36_chars(dtable_uuid)}-{page_id}')
            download_page_design_file(repo_id, dtable_uuid, page_id, is_dir, username)
            items = os.listdir(tmp_page_path)
            if 'static_image' in items:
                need_check_static = True
        else:
            download_page_design_file(repo_id, dtable_uuid, page_id, is_dir, username)
            tmp_page_path = os.path.join('/tmp/dtable-io', 'page-design', f'{uuid_str_to_36_chars(dtable_uuid)}-{page_id}.json')
    except Exception as e:
        if is_dir:
            tmp_page_path = os.path.join('/tmp/dtable-io', 'page-design', f'{uuid_str_to_36_chars(dtable_uuid)}-{page_id}')
            clear_tmp_dir(tmp_page_path)
            clear_tmp_file(tmp_page_path + '.zip')
        else:
            tmp_page_path = os.path.join('/tmp/dtable-io', 'page-design', f'{uuid_str_to_36_chars(dtable_uuid)}-{page_id}.json')
            clear_tmp_file(tmp_page_path)
        raise e
    finally:
        if is_dir:
            try:
                seafile_tmp_file = f'/asset/{uuid_str_to_36_chars(dtable_uuid)}/page-design/{uuid_str_to_36_chars(dtable_uuid)}-{page_id}.zip'
                if seafile_api.get_file_id_by_path(repo_id, seafile_tmp_file):
                    seafile_api.del_file(repo_id, os.path.dirname(seafile_tmp_file), os.path.basename(seafile_tmp_file), username)
            except Exception as e:
                dtable_io_logger.exception('delete repo: %s temp zip file: %s error: %s', repo_id, tmp_page_path, e)
        else:
            try:
                seafile_tmp_file = f'/asset/{uuid_str_to_36_chars(dtable_uuid)}/page-design/{uuid_str_to_36_chars(dtable_uuid)}-{page_id}.json'
                if seafile_api.get_file_id_by_path(repo_id, seafile_tmp_file):
                    seafile_api.del_file(repo_id, os.path.dirname(seafile_tmp_file), os.path.basename(seafile_tmp_file), username)
            except Exception as e:
                dtable_io_logger.exception('delete repo: %s temp zip file: %s error: %s', repo_id, tmp_page_path, e)

    if not os.path.exists(tmp_page_path):
        return

    try:
        if is_dir:
            # update content and save to file
            tmp_content_file = os.path.join(tmp_page_path, f'{page_id}.json')
            update_page_design_content_to_path(workspace_id, dtable_uuid, page_id, tmp_content_file, need_check_static)
        else:
            update_page_design_content_to_path(workspace_id, dtable_uuid, page_id, tmp_page_path, need_check_static)
        # upload
        upload_page_design(repo_id, dtable_uuid, page_id, tmp_page_path, is_dir, username)
    except Exception as e:
        raise e
    finally:
        if is_dir:
            clear_tmp_dir(tmp_page_path)
            clear_tmp_file(tmp_page_path + '.zip')
        else:
            clear_tmp_file(tmp_page_path)


def convert_app_table_page_to_excel(dtable_uuid, repo_id, table_id, username, app_name, page_name, filter_condition_groups, shown_column_keys, task_id, tasks_status_map, is_support_image):
    dtable_io_logger.info('Start export app table to excel: {}.'.format(dtable_uuid))
    try:
        export_app_table_page_to_excel(dtable_uuid, repo_id, table_id, username, app_name, page_name, filter_condition_groups, shown_column_keys, task_id, tasks_status_map, is_support_image)
    except Exception as e:
        dtable_io_logger.exception('export app table failed. ERROR: {}'.format(e))
    else:
        dtable_io_logger.info('export app app_name: %s, page_name: %s success!', app_name, page_name)


def export_document(repo_id, dtable_uuid, doc_uuid, parent_path, filename, username):
    """
    /tmp/document/<dtable_uuid>/<doc_uuid>/sdoc_asset/
                                |- images/
                                |- content.json
                                |- document_config.json
    zip /tmp/document/<dtable_uuid>/<doc_uuid>/sdoc_asset/ to /tmp/document/<dtable_uuid>/<doc_uuid>/zip_file.zip
    """

    tmp_base_dir = os.path.join('/tmp/document', dtable_uuid, doc_uuid)

    tmp_file_path = os.path.join(tmp_base_dir, 'sdoc_asset/')  # used to store asset files and json from file_server
    tmp_zip_path = os.path.join(tmp_base_dir, 'zip_file') + '.zip'  # zip path of zipped xxx.zip

    dtable_io_logger.info('Start prepare %s for export sdoc.', tmp_zip_path)

    dtable_io_logger.info('Clear tmp dirs and files before prepare.')
    # clear tmp files and dirs
    clear_tmp_dir(tmp_base_dir)
    os.makedirs(tmp_file_path, exist_ok=True)

    try:
        download_link = get_seadoc_download_link(repo_id, dtable_uuid, doc_uuid, parent_path, filename, is_inner=True)
        resp = requests.get(download_link)
        file_obj = io.BytesIO(resp.content)
        with open(os.path.join(tmp_file_path, 'content.json'), 'wb') as f:
            f.write(file_obj.read())
    except Exception as e:
        raise Exception('prepare sdoc failed. ERROR: %s', e)

    # export document config
    documents_settings = get_documents_config(repo_id, dtable_uuid, username)
    doc_info = next(filter(lambda t: t['doc_uuid'] == doc_uuid, documents_settings), {})
    if not doc_info:
        raise Exception('dtable_uuid: %s, doc_uuid: %s, document config file not found.', dtable_uuid, doc_uuid)

    doc_info.pop('doc_uuid', None)
    doc_info.pop('poster_url', None)
    doc_info.pop('table_id', None)
    doc_info.pop('view_id', None)

    with open(os.path.join(tmp_file_path, 'document_config.json'), 'wb') as f:
        f.write(json.dumps(doc_info).encode('utf-8'))

    #  get images folder, images could be empty
    base_dir = gen_seadoc_base_dir(dtable_uuid, doc_uuid)
    image_parent_path = os.path.join(base_dir + SDOC_IMAGES_DIR)

    images_dir_id = seafile_api.get_dir_id_by_path(repo_id, image_parent_path)
    if images_dir_id:
        dtable_io_logger.info('Create images folder.')
        try:
            export_sdoc_prepare_images_folder(repo_id, images_dir_id, username, tmp_file_path)
        except Exception as e:
            dtable_io_logger.warning('create images folder failed. ERROR: %s', e)

    dtable_io_logger.info('Make zip file for download...')
    try:
        tmp_zip_dir = os.path.join('/tmp/document', dtable_uuid, doc_uuid, 'zip_file')
        shutil.make_archive(tmp_zip_dir, 'zip', root_dir=tmp_file_path)
    except Exception as e:
        dtable_io_logger.error('make zip failed. ERROR: {}'.format(e))
        raise Exception('make zip failed. ERROR: {}'.format(e))
    dtable_io_logger.info('Create %s success!', tmp_zip_path)
    return tmp_zip_path


def import_document(repo_id, dtable_uuid, doc_uuid, view_id, table_id, username):
    """
    upload document to seaf-server
    """
    sdoc_file_name = doc_uuid + '.sdoc'
    relative_path = ''
    sdoc_parent_dir = gen_seadoc_base_dir(dtable_uuid, doc_uuid)
    tmp_extracted_path = os.path.join('/tmp/document', dtable_uuid, doc_uuid)

    seafile_api.mkdir_with_parents(repo_id, '/', sdoc_parent_dir[1:], username)

    obj_id = json.dumps({'parent_dir': sdoc_parent_dir})
    token = seafile_api.get_fileserver_access_token(repo_id, obj_id, 'upload', '', use_onetime=False)

    upload_link = gen_file_upload_url(token, 'upload-api')
    upload_link += '?ret-json=1&replace=1'

    new_file_path = os.path.join(sdoc_parent_dir, relative_path, sdoc_file_name)
    data = {'parent_dir': sdoc_parent_dir, 'target_file': new_file_path, 'relative_path': relative_path, 'replace': 1}

    sdoc_file_path = os.path.join(tmp_extracted_path, 'content.json')
    new_sdoc_file_path = os.path.join(tmp_extracted_path, sdoc_file_name)
    os.rename(sdoc_file_path, new_sdoc_file_path)

    files = {'file': open(new_sdoc_file_path, 'rb')}
    resp = requests.post(upload_link, files=files, data=data)
    if not resp.ok:
        raise Exception('upload sdoc file: %s failed: %s' % (sdoc_file_name, resp.text))

    # import document config
    document_config_file_path = os.path.join(tmp_extracted_path, 'document_config.json')
    with open(document_config_file_path, 'r') as f:
        document_config = f.read()
        doc_info = json.loads(document_config)
    doc_info['doc_uuid'] = doc_uuid
    doc_info['view_id'] = view_id
    doc_info['table_id'] = table_id

    documents_settings = get_documents_config(repo_id, dtable_uuid, username)
    documents_settings.append(doc_info)
    save_documents_config(repo_id, dtable_uuid, username, json.dumps(documents_settings))

    # upload sdoc images
    tmp_image_dir = os.path.join(tmp_extracted_path, 'images/')
    if os.path.exists(tmp_image_dir):
        batch_upload_sdoc_images(dtable_uuid, doc_uuid, repo_id, username, tmp_image_dir)

    # remove tmp file
    if os.path.exists(tmp_extracted_path):
        shutil.rmtree(tmp_extracted_path)
